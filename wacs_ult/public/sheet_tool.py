# coding: utf-8
import re
import string

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class SheetTool:
    @staticmethod
    def create_correct_sheet(dest_path: str, df: pd.DataFrame):
        wb = Workbook()
        ws = wb.active
        ws.title = 'HTML 語法錯誤'
        col_order = list(df.columns)

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        for col_name, dimension in zip(col_order, string.ascii_uppercase):
            col_strings = list(df[col_name].unique()).copy()
            col_strings.append(col_name)
            col_width = SheetTool.get_col_width(col_strings)
            ws.column_dimensions[dimension].width = col_width + 8

        fill = PatternFill('solid', fgColor='92d050')
        for item in ws['1:1']:
            item.fill = fill

        bian = Side(style='medium', color='000000')

        num = ws.max_row
        for i in range(num):
            for item in ws['1:' + str(num)][i]:
                item.border = Border(top=bian, bottom=bian, left=bian, right=bian)
                item.font = Font(name='Calibri', color='000000', size=11, b=False)

        wb.save(dest_path)
        wb.close()

    @staticmethod
    def create_result_sheet(dest_path: str, df: pd.DataFrame, escape_col=None):  # It's a beautiful sheet
        wb = Workbook()
        ws = wb.active
        col_order = list(df.columns)

        for term, domain_data in df.iterrows():
            name = domain_data[0]
            record = domain_data[1]
            record.assign(domain=name)
            record = record.reindex(columns=col_order)

            # init sheet
            if term == 0:
                ws.title = name
            else:
                ws = wb.create_sheet(name)

            # insert value
            for row in dataframe_to_rows(record, index=False, header=True):
                ws.append(row)

            # set column width
            for col_name, dimension in zip(col_order, string.ascii_uppercase):
                if col_name == 'info':
                    ws.column_dimensions[dimension].width = 40
                    continue

                col_strings = list(record[col_name].unique()).copy()
                col_strings.append(col_name)
                col_width = SheetTool.get_col_width(col_strings)
                ws.column_dimensions[dimension].width = col_width + 8

            fill = PatternFill('solid', fgColor='92d050')
            for item in ws['1:1']:
                item.fill = fill
            bian = Side(style='medium', color='000000')

            num = ws.max_row
            for i in range(num):
                for item in ws['1:' + str(num)][i]:
                    item.border = Border(top=bian, bottom=bian, left=bian, right=bian)
                    item.font = Font(name='Calibri', color='000000', size=11, b=False)

        wb.save(dest_path)
        wb.close()

    @staticmethod
    def get_col_width(words: list) -> float:
        for i, word in enumerate(words):
            word_len = 0
            for char in word:
                word_len += SheetTool.get_char_len(char)

            words[i] = round(word_len, 2)

        return max(words)

    @staticmethod
    def get_char_len(char: str) -> float:

        if SheetTool.is_chinese(char):
            return 0.8 * 2.2

        elif SheetTool.is_english(char):
            if char.isupper():
                return 0.8 * 2.2

            elif char.islower():
                return 0.5 * 2.2

        elif char.isnumeric():
            return 0.15 * 2.2

        else:
            if char == '.' or char == ' ':
                return 0.15 * 2.2
            else:
                return 0.5 * 2.2

    @staticmethod
    def is_chinese(char: str) -> bool:
        pattern = re.compile(u'[\u4e00-\u9fa5]+')
        check = pattern.search(char)
        if check:
            return True
        else:
            return False

    @staticmethod
    def is_english(char: str) -> bool:
        pattern = re.compile(r'[a-zA-Z]')
        check = pattern.search(char)
        if check:
            return True
        else:
            return False

    @staticmethod
    def get_domains_text(domain_list: list) -> str:
        result = ''
        if domain_list:
            temp = domain_list.copy()
            result = temp.pop(0)
            for domain in temp:
                result += ', ' + domain

        return result

    @staticmethod
    def get_word_width(words: str) -> float:
        word_len = 0
        for word in words:
            word_len += SheetTool.get_char_len(word)
        return word_len
